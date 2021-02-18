# PPT 자동화 프로그램 v03
# - v02+
# - 입력받은 xlsx 시트에서 ppt로 정보 이동을 자동화한 프로그램
# 작성자 : 오래규
# 작성일 : 2021-02-16

from openpyxl import load_workbook
import pyautogui
import pyperclip

franchise = str(input("업체명을 입력하시오 : "))

file_title = "우편표지 취합본 " + franchise + ".pptx"

################## 엑셀에서 업체 정보 추출 ################## 
wb = load_workbook("로드뷰 조사 리스트.xlsx") # 엑셀 파일 로딩
ws = wb.get_sheet_by_name(franchise) # 엑셀 파일 시트명 입력

address = []  # 업체 주소
names = []  # 지점명
posts = []  # 우편번호

for y in range(2, ws.max_row + 1):
  # 엑셀에서 주소 추출
  ads = str(ws.cell(column = 4, row = y).value)
  address.append(ads)
  # 엑셀에서 지점명 추출
  name = str(ws.cell(column = 2, row = y).value + " " + ws.cell(column = 3, row = y).value)
  names.append(name)
  # 엑셀에서 우편번호 추출
  post = str(ws.cell(column = 6, row = y).value)
  posts.append(post)

################## PPT에서 작업 ##################
ppt_file = pyautogui.getWindowsWithTitle(file_title)[0] # 양식이 저장되어 있는 PPT 파일 로딩
ppt_file.activate()
pyautogui.click(125, 250) # PPT 첫 슬라이드 클릭

# PPT에 작업 공간 생성 후 진행
for space in range(2, ws.max_row):
  pyautogui.hotkey("Ctrl", "d") # PPT 슬라이드 복제

pyautogui.sleep(0.5)

pyautogui.press("Home") # PPT 첫 슬라이드로 이동
pyautogui.click(1000, 300)

pyautogui.sleep(0.5)

# 업무 시작
for i in range(ws.max_row - 1): 
  pyautogui.press("Tab") # 주소 좌표 이동 후 입력
  pyautogui.press("F2")
  pyautogui.press("Delete")

  pyperclip.copy(address[i])  # copy를 뛰어 넘는 경우가 발생하여 2번 씩 작성
  pyperclip.copy(address[i])
  pyautogui.hotkey("Ctrl", "v")
  pyautogui.press("Enter")

  pyperclip.copy(names[i])
  pyperclip.copy(names[i])
  pyautogui.hotkey("Ctrl", "v")
  pyautogui.press("ESC")

  pyautogui.press("Tab") # 지점명 좌표 이동 후 입력
  pyautogui.press("F2")
  pyautogui.press("Delete")

  pyperclip.copy(names[i] + " 사장님 귀하")
  pyperclip.copy(names[i] + " 사장님 귀하")
  pyautogui.hotkey("Ctrl", "v")
  pyautogui.press("ESC")
  
  for j in range(8):
    pyautogui.press("Tab") # 우편번호 좌표 이동 후 입력
  pyautogui.press("F2")
  pyautogui.press("Delete")

  pyautogui.write(posts[i])
  if i is not ws.max_row - 2:
    pyautogui.press("PageDown")   